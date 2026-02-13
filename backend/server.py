from fastapi import FastAPI, APIRouter, HTTPException, Depends, UploadFile, File, Form, status
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
from fastapi.responses import StreamingResponse
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field, EmailStr, validator
from typing import List, Optional
import uuid
from datetime import datetime, timezone, timedelta
import jwt
import bcrypt
import io
from openpyxl import Workbook, load_workbook
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.units import inch
import base64

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# JWT Config
JWT_SECRET = os.environ.get('JWT_SECRET')
if not JWT_SECRET:
    raise ValueError("JWT_SECRET environment variable is required")
JWT_ALGORITHM = "HS256"
JWT_EXPIRATION_HOURS = 24

app = FastAPI(title="Dashboard Agro Mopomulo API")
api_router = APIRouter(prefix="/api")
security = HTTPBearer()

# CORS Configuration - Support for deployment
cors_origins_env = os.environ.get('CORS_ORIGINS', '*')
if cors_origins_env == '*':
    cors_origins = ["*"]
else:
    cors_origins = [origin.strip() for origin in cors_origins_env.split(',')]

app.add_middleware(
    CORSMiddleware,
    allow_origins=cors_origins,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# ============== MODELS ==============

class UserCreate(BaseModel):
    email: EmailStr
    password: str
    nama: str

class UserLogin(BaseModel):
    email: EmailStr
    password: str

class UserResponse(BaseModel):
    id: str
    email: str
    nama: str
    role: str

class OPDCreate(BaseModel):
    nama: str
    kode: Optional[str] = None
    alamat: Optional[str] = None
    jumlah_personil: Optional[int] = 0
    kategori: Optional[str] = "OPD"  # OPD, DESA, PUBLIK

class OPDUpdate(BaseModel):
    nama: Optional[str] = None
    kode: Optional[str] = None
    alamat: Optional[str] = None
    jumlah_personil: Optional[int] = None
    kategori: Optional[str] = None

class OPDResponse(BaseModel):
    id: str
    nama: str
    kode: Optional[str] = None
    alamat: Optional[str] = None
    jumlah_personil: Optional[int] = 0
    kategori: Optional[str] = "OPD"
    created_at: str

# Model untuk lokasi tanam (per titik)
class LokasiTanam(BaseModel):
    lokasi_tanam: str
    titik_lokasi: Optional[str] = None
    bukti_url: Optional[str] = None

class PartisipasiCreate(BaseModel):
    email: Optional[str] = None
    nama_lengkap: str
    nip: Optional[str] = None
    opd_id: str
    alamat: Optional[str] = None
    nomor_whatsapp: Optional[str] = None
    jumlah_pohon: int
    jenis_pohon: str
    sumber_bibit: str
    # Support both single lokasi (backward compatible) and array of lokasi
    lokasi_tanam: Optional[str] = None
    titik_lokasi: Optional[str] = None
    bukti_url: Optional[str] = None
    # Array of multiple locations
    lokasi_list: Optional[List[LokasiTanam]] = None
    
    @validator('email', pre=True)
    def validate_email(cls, v):
        if v is None or v == '':
            return None
        # Basic email validation
        import re
        if v and not re.match(r'^[\w\.-]+@[\w\.-]+\.\w+$', v):
            raise ValueError('Format email tidak valid')
        return v

class PartisipasiUpdate(BaseModel):
    email: Optional[str] = None
    nama_lengkap: Optional[str] = None
    nip: Optional[str] = None
    opd_id: Optional[str] = None
    alamat: Optional[str] = None
    nomor_whatsapp: Optional[str] = None
    jumlah_pohon: Optional[int] = None
    jenis_pohon: Optional[str] = None
    sumber_bibit: Optional[str] = None
    lokasi_tanam: Optional[str] = None
    titik_lokasi: Optional[str] = None
    bukti_url: Optional[str] = None
    status: Optional[str] = None
    # Array of multiple locations
    lokasi_list: Optional[List[LokasiTanam]] = None

class PartisipasiResponse(BaseModel):
    id: str
    email: Optional[str] = None
    nama_lengkap: str
    nip: Optional[str] = None
    opd_id: str
    opd_nama: Optional[str] = None
    alamat: Optional[str] = None
    nomor_whatsapp: Optional[str] = None
    jumlah_pohon: int
    jenis_pohon: str
    sumber_bibit: Optional[str] = None
    # Single lokasi (for backward compatibility)
    lokasi_tanam: Optional[str] = None
    titik_lokasi: Optional[str] = None
    bukti_url: Optional[str] = None
    # Array of multiple locations
    lokasi_list: Optional[List[dict]] = None
    status: Optional[str] = None
    created_at: str

class SettingsUpdate(BaseModel):
    logo_url: Optional[str] = None
    hero_title: Optional[str] = None
    hero_subtitle: Optional[str] = None
    hero_image_url: Optional[str] = None
    tentang_title: Optional[str] = None
    tentang_content: Optional[str] = None
    tentang_visi: Optional[str] = None
    tentang_misi: Optional[str] = None
    berita_popup_interval: Optional[int] = None  # dalam detik

class SettingsResponse(BaseModel):
    id: str
    logo_url: Optional[str] = None
    hero_title: str
    hero_subtitle: str
    hero_image_url: Optional[str] = None
    tentang_title: Optional[str] = None
    tentang_content: Optional[str] = None
    tentang_visi: Optional[str] = None
    tentang_misi: Optional[str] = None
    berita_popup_interval: Optional[int] = 5  # default 5 detik

class GalleryCreate(BaseModel):
    title: str
    image_url: str
    description: Optional[str] = None

class GalleryResponse(BaseModel):
    id: str
    title: str
    image_url: str
    description: Optional[str] = None
    created_at: str

class EdukasiCreate(BaseModel):
    judul: str
    konten: str
    gambar_url: Optional[str] = None

class EdukasiUpdate(BaseModel):
    judul: Optional[str] = None
    konten: Optional[str] = None
    gambar_url: Optional[str] = None

class EdukasiResponse(BaseModel):
    id: str
    judul: str
    konten: str
    gambar_url: Optional[str] = None
    created_at: str

# ============== AGENDA MODELS ==============

class AgendaCreate(BaseModel):
    nama_kegiatan: str
    hari: str
    tanggal: str
    lokasi_kecamatan: str
    lokasi_desa: str
    deskripsi: Optional[str] = None

class AgendaUpdate(BaseModel):
    nama_kegiatan: Optional[str] = None
    hari: Optional[str] = None
    tanggal: Optional[str] = None
    lokasi_kecamatan: Optional[str] = None
    lokasi_desa: Optional[str] = None
    deskripsi: Optional[str] = None
    status: Optional[str] = None

class AgendaResponse(BaseModel):
    id: str
    nama_kegiatan: str
    hari: str
    tanggal: str
    lokasi_kecamatan: str
    lokasi_desa: str
    deskripsi: Optional[str] = None
    status: str
    created_at: str

# ============== BERITA MODELS ==============

class BeritaCreate(BaseModel):
    judul: str
    deskripsi_singkat: str
    link_berita: str  # URL ke halaman berita eksternal
    isi_berita: Optional[str] = None  # Deprecated, untuk backward compatibility
    gambar_url: Optional[str] = None
    gambar_type: Optional[str] = "link"  # "link" atau "file"

class BeritaUpdate(BaseModel):
    judul: Optional[str] = None
    deskripsi_singkat: Optional[str] = None
    link_berita: Optional[str] = None
    isi_berita: Optional[str] = None
    gambar_url: Optional[str] = None
    gambar_type: Optional[str] = None
    is_active: Optional[bool] = None

class BeritaResponse(BaseModel):
    id: str
    judul: str
    deskripsi_singkat: str
    link_berita: Optional[str] = None
    isi_berita: Optional[str] = None  # Untuk backward compatibility
    gambar_url: Optional[str] = None
    gambar_type: str
    is_active: bool
    created_at: str

# ============== KONTAK WHATSAPP MODELS ==============

class KontakWhatsAppCreate(BaseModel):
    nomor_whatsapp: str
    pesan_default: Optional[str] = None
    
    @validator('nomor_whatsapp')
    def normalize_phone(cls, v):
        if not v:
            raise ValueError('Nomor WhatsApp wajib diisi')
        # Hapus karakter non-digit kecuali +
        cleaned = ''.join(c for c in v if c.isdigit() or c == '+')
        # Hapus + di awal
        cleaned = cleaned.lstrip('+')
        # Konversi format 08 ke 628
        if cleaned.startswith('08'):
            cleaned = '62' + cleaned[1:]
        # Pastikan dimulai dengan 62
        elif not cleaned.startswith('62'):
            cleaned = '62' + cleaned
        # Validasi panjang (minimal 10 digit setelah 62)
        if len(cleaned) < 10:
            raise ValueError('Nomor WhatsApp tidak valid')
        return cleaned

class KontakWhatsAppResponse(BaseModel):
    nomor_whatsapp: str
    pesan_default: Optional[str] = None
    updated_at: Optional[str] = None

# ============== AUTH HELPERS ==============

def hash_password(password: str) -> str:
    return bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')

def verify_password(password: str, hashed: str) -> bool:
    return bcrypt.checkpw(password.encode('utf-8'), hashed.encode('utf-8'))

def create_token(user_id: str, email: str, role: str) -> str:
    payload = {
        "user_id": user_id,
        "email": email,
        "role": role,
        "exp": datetime.now(timezone.utc) + timedelta(hours=JWT_EXPIRATION_HOURS)
    }
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGORITHM)

async def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)):
    try:
        payload = jwt.decode(credentials.credentials, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        return payload
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token telah kadaluarsa")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Token tidak valid")

# ============== AUTH ENDPOINTS ==============

@api_router.post("/auth/register", response_model=dict)
async def register(user: UserCreate):
    existing = await db.users.find_one({"email": user.email})
    if existing:
        raise HTTPException(status_code=400, detail="Email sudah terdaftar")
    
    user_id = str(uuid.uuid4())
    user_doc = {
        "id": user_id,
        "email": user.email,
        "password": hash_password(user.password),
        "nama": user.nama,
        "role": "admin",
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.users.insert_one(user_doc)
    token = create_token(user_id, user.email, "admin")
    return {"token": token, "user": {"id": user_id, "email": user.email, "nama": user.nama, "role": "admin"}}

@api_router.post("/auth/login", response_model=dict)
async def login(user: UserLogin):
    existing = await db.users.find_one({"email": user.email}, {"_id": 0})
    if not existing or not verify_password(user.password, existing["password"]):
        raise HTTPException(status_code=401, detail="Email atau password salah")
    
    token = create_token(existing["id"], existing["email"], existing["role"])
    return {"token": token, "user": {"id": existing["id"], "email": existing["email"], "nama": existing["nama"], "role": existing["role"]}}

@api_router.get("/auth/me", response_model=UserResponse)
async def get_me(current_user: dict = Depends(get_current_user)):
    user = await db.users.find_one({"id": current_user["user_id"]}, {"_id": 0, "password": 0})
    if not user:
        raise HTTPException(status_code=404, detail="User tidak ditemukan")
    return user

# ============== OPD ENDPOINTS ==============

@api_router.get("/opd", response_model=List[OPDResponse])
async def get_all_opd():
    opd_list = await db.opd.find({}, {"_id": 0}).to_list(1000)
    return opd_list

@api_router.get("/opd/{opd_id}", response_model=OPDResponse)
async def get_opd(opd_id: str):
    opd = await db.opd.find_one({"id": opd_id}, {"_id": 0})
    if not opd:
        raise HTTPException(status_code=404, detail="OPD tidak ditemukan")
    return opd

@api_router.post("/opd", response_model=OPDResponse)
async def create_opd(opd: OPDCreate, current_user: dict = Depends(get_current_user)):
    opd_id = str(uuid.uuid4())
    opd_doc = {
        "id": opd_id,
        "nama": opd.nama,
        "kode": opd.kode,
        "alamat": opd.alamat,
        "jumlah_personil": opd.jumlah_personil or 0,
        "kategori": opd.kategori or "OPD",
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.opd.insert_one(opd_doc)
    return {**opd_doc}

@api_router.put("/opd/{opd_id}", response_model=OPDResponse)
async def update_opd(opd_id: str, opd: OPDUpdate, current_user: dict = Depends(get_current_user)):
    update_data = {k: v for k, v in opd.model_dump().items() if v is not None}
    if not update_data:
        raise HTTPException(status_code=400, detail="Tidak ada data untuk diupdate")
    
    result = await db.opd.update_one({"id": opd_id}, {"$set": update_data})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="OPD tidak ditemukan")
    
    updated = await db.opd.find_one({"id": opd_id}, {"_id": 0})
    return updated

@api_router.delete("/opd/{opd_id}")
async def delete_opd(opd_id: str, current_user: dict = Depends(get_current_user)):
    result = await db.opd.delete_one({"id": opd_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="OPD tidak ditemukan")
    return {"message": "OPD berhasil dihapus"}

@api_router.post("/opd/import")
async def import_opd_excel(
    file: UploadFile = File(...),
    kategori: str = Form(...),
    current_user: dict = Depends(get_current_user)
):
    """Import OPD data from Excel file"""
    import pandas as pd
    import io
    
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="File harus berformat Excel (.xlsx atau .xls)")
    
    try:
        contents = await file.read()
        df = pd.read_excel(io.BytesIO(contents))
        
        # Normalize column names (lowercase and strip whitespace)
        df.columns = df.columns.str.lower().str.strip()
        
        # Check for required column 'nama'
        if 'nama' not in df.columns:
            raise HTTPException(status_code=400, detail="Kolom 'Nama' wajib ada dalam file Excel")
        
        imported_count = 0
        skipped_count = 0
        
        for _, row in df.iterrows():
            nama = str(row.get('nama', '')).strip()
            if not nama or nama == 'nan':
                skipped_count += 1
                continue
                
            # Check if OPD with same name and kategori already exists
            existing = await db.opd.find_one({"nama": nama, "kategori": kategori})
            if existing:
                skipped_count += 1
                continue
            
            opd_doc = {
                "id": str(uuid.uuid4()),
                "nama": nama,
                "kode": str(row.get('kode', '')).strip() if pd.notna(row.get('kode')) else '',
                "alamat": str(row.get('alamat', '')).strip() if pd.notna(row.get('alamat')) else '',
                "jumlah_personil": int(row.get('jumlah_personil', 0)) if pd.notna(row.get('jumlah_personil')) else 0,
                "kategori": kategori,
                "created_at": datetime.now(timezone.utc).isoformat()
            }
            
            await db.opd.insert_one(opd_doc)
            imported_count += 1
        
        return {
            "message": f"Import berhasil! {imported_count} data ditambahkan, {skipped_count} data dilewati (duplikat/kosong)",
            "imported": imported_count,
            "skipped": skipped_count
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Gagal import: {str(e)}")

# ============== PARTISIPASI ENDPOINTS ==============

@api_router.get("/partisipasi", response_model=List[PartisipasiResponse])
async def get_all_partisipasi():
    partisipasi_list = await db.partisipasi.find({}, {"_id": 0}).to_list(10000)
    # Fetch all OPDs once to avoid N+1 query
    opd_list = await db.opd.find({}, {"_id": 0, "id": 1, "nama": 1}).to_list(1000)
    opd_map = {o["id"]: o["nama"] for o in opd_list}
    # Enrich with OPD nama using the map
    for p in partisipasi_list:
        p["opd_nama"] = opd_map.get(p.get("opd_id"), "Unknown")
    return partisipasi_list

@api_router.get("/partisipasi/{partisipasi_id}", response_model=PartisipasiResponse)
async def get_partisipasi(partisipasi_id: str):
    p = await db.partisipasi.find_one({"id": partisipasi_id}, {"_id": 0})
    if not p:
        raise HTTPException(status_code=404, detail="Partisipasi tidak ditemukan")
    opd = await db.opd.find_one({"id": p.get("opd_id")}, {"_id": 0})
    p["opd_nama"] = opd["nama"] if opd else "Unknown"
    return p

@api_router.post("/partisipasi", response_model=PartisipasiResponse)
async def create_partisipasi(data: PartisipasiCreate):
    # Verify OPD exists
    opd = await db.opd.find_one({"id": data.opd_id}, {"_id": 0})
    if not opd:
        raise HTTPException(status_code=400, detail="OPD tidak ditemukan")
    
    partisipasi_id = str(uuid.uuid4())
    
    # Handle lokasi_list (array of locations)
    lokasi_list = []
    if data.lokasi_list and len(data.lokasi_list) > 0:
        # Use lokasi_list array
        for loc in data.lokasi_list:
            lokasi_list.append({
                "lokasi_tanam": loc.lokasi_tanam,
                "titik_lokasi": loc.titik_lokasi,
                "bukti_url": loc.bukti_url
            })
        # Set primary lokasi from first item for backward compatibility
        primary_lokasi = data.lokasi_list[0]
        lokasi_tanam = primary_lokasi.lokasi_tanam
        titik_lokasi = primary_lokasi.titik_lokasi
        bukti_url = primary_lokasi.bukti_url
    else:
        # Single lokasi (backward compatible)
        lokasi_tanam = data.lokasi_tanam or ""
        titik_lokasi = data.titik_lokasi
        bukti_url = data.bukti_url
        if lokasi_tanam:
            lokasi_list.append({
                "lokasi_tanam": lokasi_tanam,
                "titik_lokasi": titik_lokasi,
                "bukti_url": bukti_url
            })
    
    doc = {
        "id": partisipasi_id,
        "email": data.email,
        "nama_lengkap": data.nama_lengkap,
        "nip": data.nip,
        "opd_id": data.opd_id,
        "alamat": data.alamat,
        "nomor_whatsapp": data.nomor_whatsapp,
        "jumlah_pohon": data.jumlah_pohon,
        "jenis_pohon": data.jenis_pohon,
        "sumber_bibit": data.sumber_bibit,
        "lokasi_tanam": lokasi_tanam,
        "titik_lokasi": titik_lokasi,
        "bukti_url": bukti_url,
        "lokasi_list": lokasi_list,
        "status": "pending",
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.partisipasi.insert_one(doc)
    return {**doc, "opd_nama": opd["nama"]}

@api_router.put("/partisipasi/{partisipasi_id}", response_model=PartisipasiResponse)
async def update_partisipasi(partisipasi_id: str, data: PartisipasiUpdate, current_user: dict = Depends(get_current_user)):
    update_data = {k: v for k, v in data.model_dump().items() if v is not None}
    if not update_data:
        raise HTTPException(status_code=400, detail="Tidak ada data untuk diupdate")
    
    result = await db.partisipasi.update_one({"id": partisipasi_id}, {"$set": update_data})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Partisipasi tidak ditemukan")
    
    updated = await db.partisipasi.find_one({"id": partisipasi_id}, {"_id": 0})
    opd = await db.opd.find_one({"id": updated.get("opd_id")}, {"_id": 0})
    updated["opd_nama"] = opd["nama"] if opd else "Unknown"
    return updated

@api_router.delete("/partisipasi/{partisipasi_id}")
async def delete_partisipasi(partisipasi_id: str, current_user: dict = Depends(get_current_user)):
    result = await db.partisipasi.delete_one({"id": partisipasi_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Partisipasi tidak ditemukan")
    return {"message": "Partisipasi berhasil dihapus"}

# ============== SETTINGS ENDPOINTS ==============

@api_router.get("/settings", response_model=SettingsResponse)
async def get_settings():
    settings = await db.settings.find_one({}, {"_id": 0})
    if not settings:
        # Return defaults
        default_settings = {
            "id": str(uuid.uuid4()),
            "logo_url": None,
            "hero_title": "Gerakan Agro Mopomulo",
            "hero_subtitle": "Satu Orang Sepuluh Pohon untuk Masa Depan Daerah",
            "hero_image_url": "https://images.unsplash.com/photo-1765333534690-ad3a985e7c42?crop=entropy&cs=srgb&fm=jpg&ixid=M3w3NDQ2NDJ8MHwxfHNlYXJjaHwxfHxsdXNoJTIwZ3JlZW4lMjBmb3Jlc3QlMjBsYW5kc2NhcGUlMjBpbmRvbmVzaWF8ZW58MHx8fHwxNzY4NDQ1ODE1fDA&ixlib=rb-4.1.0&q=85",
            "tentang_title": "Program Agro Mopomulo",
            "tentang_content": "Mopomulo berasal dari bahasa Gorontalo yang berarti \"menanam\". Program Agro Mopomulo adalah inisiatif Pemerintah Kabupaten Gorontalo Utara untuk meningkatkan kesadaran dan partisipasi masyarakat dalam pelestarian lingkungan.\n\nDengan konsep \"Satu Orang Sepuluh Pohon\", program ini menargetkan setiap ASN dan warga untuk berkontribusi menanam minimal 10 pohon, baik pohon produktif maupun pohon pelindung.",
            "tentang_visi": "Mewujudkan Kabupaten Gorontalo Utara sebagai daerah yang hijau, asri, dan berkelanjutan dengan partisipasi aktif seluruh lapisan masyarakat dalam pelestarian lingkungan.",
            "tentang_misi": "- Meningkatkan kesadaran lingkungan masyarakat\n- Memperluas area hijau di seluruh wilayah\n- Mendukung ketahanan pangan daerah\n- Membangun budaya peduli lingkungan",
            "berita_popup_interval": 5
        }
        await db.settings.insert_one(default_settings)
        return default_settings
    # Ensure berita_popup_interval exists
    if "berita_popup_interval" not in settings:
        settings["berita_popup_interval"] = 5
    return settings

@api_router.put("/settings", response_model=SettingsResponse)
async def update_settings(data: SettingsUpdate, current_user: dict = Depends(get_current_user)):
    update_data = {k: v for k, v in data.model_dump().items() if v is not None}
    
    existing = await db.settings.find_one({}, {"_id": 0})
    if existing:
        await db.settings.update_one({"id": existing["id"]}, {"$set": update_data})
        updated = await db.settings.find_one({"id": existing["id"]}, {"_id": 0})
        return updated
    else:
        settings_id = str(uuid.uuid4())
        new_settings = {
            "id": settings_id,
            "logo_url": data.logo_url,
            "hero_title": data.hero_title or "Gerakan Agro Mopomulo",
            "hero_subtitle": data.hero_subtitle or "Satu Orang Sepuluh Pohon untuk Masa Depan Daerah",
            "hero_image_url": data.hero_image_url,
            "tentang_title": data.tentang_title or "Program Agro Mopomulo",
            "tentang_content": data.tentang_content,
            "tentang_visi": data.tentang_visi,
            "tentang_misi": data.tentang_misi,
            "berita_popup_interval": data.berita_popup_interval or 5
        }
        await db.settings.insert_one(new_settings)
        return new_settings

@api_router.post("/upload/image")
async def upload_image(file: UploadFile = File(...)):
    # Validate file size (max 2MB)
    MAX_FILE_SIZE = 2 * 1024 * 1024  # 2MB
    contents = await file.read()
    
    if len(contents) > MAX_FILE_SIZE:
        raise HTTPException(status_code=400, detail="Ukuran file maksimal 2MB")
    
    # Validate file type
    if not file.content_type or not file.content_type.startswith('image/'):
        raise HTTPException(status_code=400, detail="File harus berupa gambar")
    
    base64_data = base64.b64encode(contents).decode('utf-8')
    content_type = file.content_type or 'image/png'
    data_url = f"data:{content_type};base64,{base64_data}"
    return {"url": data_url}

@api_router.post("/settings/upload-logo")
async def upload_logo(file: UploadFile = File(...), current_user: dict = Depends(get_current_user)):
    # Validate file size (max 2MB)
    MAX_FILE_SIZE = 2 * 1024 * 1024  # 2MB
    contents = await file.read()
    
    if len(contents) > MAX_FILE_SIZE:
        raise HTTPException(status_code=400, detail="Ukuran file maksimal 2MB")
    
    base64_data = base64.b64encode(contents).decode('utf-8')
    content_type = file.content_type or 'image/png'
    data_url = f"data:{content_type};base64,{base64_data}"
    
    existing = await db.settings.find_one({}, {"_id": 0})
    if existing:
        await db.settings.update_one({"id": existing["id"]}, {"$set": {"logo_url": data_url}})
    else:
        settings_id = str(uuid.uuid4())
        new_settings = {
            "id": settings_id,
            "logo_url": data_url,
            "hero_title": "Gerakan Agro Mopomulo",
            "hero_subtitle": "Satu Orang Sepuluh Pohon untuk Masa Depan Daerah",
            "hero_image_url": None
        }
        await db.settings.insert_one(new_settings)
    
    return {"logo_url": data_url}

# ============== GALLERY ENDPOINTS ==============

@api_router.get("/gallery", response_model=List[GalleryResponse])
async def get_all_gallery():
    items = await db.gallery.find({}, {"_id": 0}).to_list(1000)
    return items

@api_router.post("/gallery", response_model=GalleryResponse)
async def create_gallery(data: GalleryCreate, current_user: dict = Depends(get_current_user)):
    gallery_id = str(uuid.uuid4())
    doc = {
        "id": gallery_id,
        **data.model_dump(),
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.gallery.insert_one(doc)
    return doc

@api_router.delete("/gallery/{gallery_id}")
async def delete_gallery(gallery_id: str, current_user: dict = Depends(get_current_user)):
    result = await db.gallery.delete_one({"id": gallery_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Item galeri tidak ditemukan")
    return {"message": "Item galeri berhasil dihapus"}

# ============== EDUKASI ENDPOINTS ==============

@api_router.get("/edukasi", response_model=List[EdukasiResponse])
async def get_all_edukasi():
    items = await db.edukasi.find({}, {"_id": 0}).to_list(1000)
    return items

@api_router.post("/edukasi", response_model=EdukasiResponse)
async def create_edukasi(data: EdukasiCreate, current_user: dict = Depends(get_current_user)):
    edukasi_id = str(uuid.uuid4())
    doc = {
        "id": edukasi_id,
        **data.model_dump(),
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.edukasi.insert_one(doc)
    return doc

@api_router.put("/edukasi/{edukasi_id}", response_model=EdukasiResponse)
async def update_edukasi(edukasi_id: str, data: EdukasiUpdate, current_user: dict = Depends(get_current_user)):
    update_data = {k: v for k, v in data.model_dump().items() if v is not None}
    if not update_data:
        raise HTTPException(status_code=400, detail="Tidak ada data untuk diupdate")
    result = await db.edukasi.update_one({"id": edukasi_id}, {"$set": update_data})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Edukasi tidak ditemukan")
    updated = await db.edukasi.find_one({"id": edukasi_id}, {"_id": 0})
    return updated

@api_router.delete("/edukasi/{edukasi_id}")
async def delete_edukasi(edukasi_id: str, current_user: dict = Depends(get_current_user)):
    result = await db.edukasi.delete_one({"id": edukasi_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Edukasi tidak ditemukan")
    return {"message": "Edukasi berhasil dihapus"}

# ============== AGENDA ENDPOINTS ==============

@api_router.get("/agenda", response_model=List[AgendaResponse])
async def get_all_agenda():
    items = await db.agenda.find({}, {"_id": 0}).sort("tanggal", 1).to_list(1000)
    return items

@api_router.get("/agenda/upcoming", response_model=List[AgendaResponse])
async def get_upcoming_agenda():
    """Get upcoming agenda (status = upcoming or ongoing)"""
    items = await db.agenda.find(
        {"status": {"$in": ["upcoming", "ongoing"]}}, 
        {"_id": 0}
    ).sort("tanggal", 1).to_list(10)
    return items

@api_router.post("/agenda", response_model=AgendaResponse)
async def create_agenda(data: AgendaCreate, current_user: dict = Depends(get_current_user)):
    agenda_id = str(uuid.uuid4())
    doc = {
        "id": agenda_id,
        **data.model_dump(),
        "status": "upcoming",  # upcoming, ongoing, completed
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.agenda.insert_one(doc)
    return doc

@api_router.put("/agenda/{agenda_id}", response_model=AgendaResponse)
async def update_agenda(agenda_id: str, data: AgendaUpdate, current_user: dict = Depends(get_current_user)):
    update_data = {k: v for k, v in data.model_dump().items() if v is not None}
    if not update_data:
        raise HTTPException(status_code=400, detail="Tidak ada data untuk diupdate")
    result = await db.agenda.update_one({"id": agenda_id}, {"$set": update_data})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Agenda tidak ditemukan")
    updated = await db.agenda.find_one({"id": agenda_id}, {"_id": 0})
    return updated

@api_router.delete("/agenda/{agenda_id}")
async def delete_agenda(agenda_id: str, current_user: dict = Depends(get_current_user)):
    result = await db.agenda.delete_one({"id": agenda_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Agenda tidak ditemukan")
    return {"message": "Agenda berhasil dihapus"}

# ============== BERITA ENDPOINTS ==============

@api_router.get("/berita", response_model=List[BeritaResponse])
async def get_all_berita():
    items = await db.berita.find({}, {"_id": 0}).sort("created_at", -1).to_list(1000)
    return items

@api_router.get("/berita/active", response_model=List[BeritaResponse])
async def get_active_berita():
    """Get active news for popup"""
    items = await db.berita.find(
        {"is_active": True}, 
        {"_id": 0}
    ).sort("created_at", -1).to_list(10)
    return items

@api_router.get("/berita/{berita_id}", response_model=BeritaResponse)
async def get_berita_by_id(berita_id: str):
    berita = await db.berita.find_one({"id": berita_id}, {"_id": 0})
    if not berita:
        raise HTTPException(status_code=404, detail="Berita tidak ditemukan")
    return berita

@api_router.post("/berita", response_model=BeritaResponse)
async def create_berita(data: BeritaCreate, current_user: dict = Depends(get_current_user)):
    berita_id = str(uuid.uuid4())
    doc = {
        "id": berita_id,
        **data.model_dump(),
        "is_active": True,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.berita.insert_one(doc)
    return doc

@api_router.put("/berita/{berita_id}", response_model=BeritaResponse)
async def update_berita(berita_id: str, data: BeritaUpdate, current_user: dict = Depends(get_current_user)):
    update_data = {k: v for k, v in data.model_dump().items() if v is not None}
    if not update_data:
        raise HTTPException(status_code=400, detail="Tidak ada data untuk diupdate")
    result = await db.berita.update_one({"id": berita_id}, {"$set": update_data})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Berita tidak ditemukan")
    updated = await db.berita.find_one({"id": berita_id}, {"_id": 0})
    return updated

@api_router.delete("/berita/{berita_id}")
async def delete_berita(berita_id: str, current_user: dict = Depends(get_current_user)):
    result = await db.berita.delete_one({"id": berita_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Berita tidak ditemukan")
    return {"message": "Berita berhasil dihapus"}

# ============== STATS ENDPOINTS ==============

@api_router.get("/stats")
async def get_stats():
    total_pohon = 0
    total_partisipan = 0
    
    # Ambil SEMUA data partisipasi (tidak filter by status)
    partisipasi_list = await db.partisipasi.find({}, {"_id": 0}).to_list(10000)
    total_partisipan = len(partisipasi_list)
    total_pohon = sum(p.get("jumlah_pohon", 0) for p in partisipasi_list)
    
    # Stats per OPD
    opd_stats = {}
    for p in partisipasi_list:
        opd_id = p.get("opd_id")
        if opd_id not in opd_stats:
            opd_stats[opd_id] = {"jumlah_pohon": 0, "jumlah_partisipan": 0}
        opd_stats[opd_id]["jumlah_pohon"] += p.get("jumlah_pohon", 0)
        opd_stats[opd_id]["jumlah_partisipan"] += 1
    
    # Enrich with OPD names
    opd_list = await db.opd.find({}, {"_id": 0}).to_list(1000)
    opd_map = {o["id"]: o["nama"] for o in opd_list}
    
    opd_stats_list = []
    for opd_id, stats in opd_stats.items():
        opd_stats_list.append({
            "opd_id": opd_id,
            "opd_nama": opd_map.get(opd_id, "Unknown"),
            "jumlah_pohon": stats["jumlah_pohon"],
            "jumlah_partisipan": stats["jumlah_partisipan"]
        })
    
    # Stats per jenis pohon
    jenis_pohon_stats = {}
    for p in partisipasi_list:
        jenis = p.get("jenis_pohon", "Lainnya")
        if jenis not in jenis_pohon_stats:
            jenis_pohon_stats[jenis] = 0
        jenis_pohon_stats[jenis] += p.get("jumlah_pohon", 0)
    
    jenis_pohon_list = [{"jenis": k, "jumlah": v} for k, v in jenis_pohon_stats.items()]
    
    # Lokasi tanam stats - hitung dari lokasi_list atau single lokasi
    # Perhatian: Satu partisipan dengan multiple lokasi tetap dihitung sebagai 1 orang per partisipan
    lokasi_stats = {}
    lokasi_partisipan_ids = {}  # Track partisipan IDs per lokasi untuk mencegah double counting
    total_lokasi = 0
    
    for p in partisipasi_list:
        partisipan_id = p.get("id", str(id(p)))  # Unique identifier untuk partisipan ini
        
        # Prioritaskan lokasi_list jika ada
        lokasi_list = p.get("lokasi_list", [])
        if lokasi_list and len(lokasi_list) > 0:
            num_lokasi = len(lokasi_list)
            pohon_per_lokasi = p.get("jumlah_pohon", 0) // num_lokasi if num_lokasi > 0 else 0
            
            for loc in lokasi_list:
                lokasi = loc.get("lokasi_tanam", "Tidak diketahui")
                if lokasi and lokasi.strip():
                    if lokasi not in lokasi_stats:
                        lokasi_stats[lokasi] = {"jumlah_pohon": 0, "jumlah_partisipan": 0}
                        lokasi_partisipan_ids[lokasi] = set()
                    
                    # Tambah pohon
                    lokasi_stats[lokasi]["jumlah_pohon"] += pohon_per_lokasi
                    total_lokasi += 1
                    
                    # Hanya tambah partisipan jika belum dihitung untuk lokasi ini
                    if partisipan_id not in lokasi_partisipan_ids[lokasi]:
                        lokasi_stats[lokasi]["jumlah_partisipan"] += 1
                        lokasi_partisipan_ids[lokasi].add(partisipan_id)
        else:
            # Fallback ke single lokasi
            lokasi = p.get("lokasi_tanam", "Tidak diketahui")
            if lokasi and lokasi.strip():
                if lokasi not in lokasi_stats:
                    lokasi_stats[lokasi] = {"jumlah_pohon": 0, "jumlah_partisipan": 0}
                    lokasi_partisipan_ids[lokasi] = set()
                
                lokasi_stats[lokasi]["jumlah_pohon"] += p.get("jumlah_pohon", 0)
                total_lokasi += 1
                
                # Hanya tambah partisipan jika belum dihitung
                if partisipan_id not in lokasi_partisipan_ids[lokasi]:
                    lokasi_stats[lokasi]["jumlah_partisipan"] += 1
                    lokasi_partisipan_ids[lokasi].add(partisipan_id)
    
    lokasi_list_result = [{"lokasi": k, **v} for k, v in lokasi_stats.items()]
    
    return {
        "total_pohon": total_pohon,
        "total_partisipan": total_partisipan,
        "total_opd": len(opd_list),
        "total_lokasi": total_lokasi,
        "opd_stats": sorted(opd_stats_list, key=lambda x: x["jumlah_pohon"], reverse=True),
        "jenis_pohon_stats": sorted(jenis_pohon_list, key=lambda x: x["jumlah"], reverse=True),
        "lokasi_stats": sorted(lokasi_list_result, key=lambda x: x["jumlah_pohon"], reverse=True)
    }

@api_router.get("/progress")
async def get_progress():
    """Get progress per OPD based on formula: target = 10 trees × personnel"""
    # Get all OPDs with personnel count
    opd_list = await db.opd.find({}, {"_id": 0}).to_list(1000)
    
    # Get ALL partisipasi (tidak filter by status)
    partisipasi_list = await db.partisipasi.find({}, {"_id": 0}).to_list(10000)
    
    # Calculate planted trees per OPD
    opd_planted = {}
    for p in partisipasi_list:
        opd_id = p.get("opd_id")
        if opd_id not in opd_planted:
            opd_planted[opd_id] = 0
        opd_planted[opd_id] += p.get("jumlah_pohon", 0)
    
    # Build progress list
    progress_list = []
    total_target = 0
    total_planted = 0
    
    for opd in opd_list:
        jumlah_personil = opd.get("jumlah_personil", 0) or 0
        target = jumlah_personil * 10  # 1 orang = 10 pohon
        planted = opd_planted.get(opd["id"], 0)
        progress_pct = round((planted / target * 100), 1) if target > 0 else 0
        
        total_target += target
        total_planted += planted
        
        progress_list.append({
            "opd_id": opd["id"],
            "opd_nama": opd["nama"],
            "kategori": opd.get("kategori", "OPD"),
            "jumlah_personil": jumlah_personil,
            "target_pohon": target,
            "pohon_tertanam": planted,
            "progress_persen": min(progress_pct, 100)  # Cap at 100%
        })
    
    # Sort by progress percentage descending
    progress_list = sorted(progress_list, key=lambda x: x["progress_persen"], reverse=True)
    
    # Calculate overall progress
    overall_progress = round((total_planted / total_target * 100), 1) if total_target > 0 else 0
    
    return {
        "progress_list": progress_list,
        "summary": {
            "total_personil": sum(o.get("jumlah_personil", 0) or 0 for o in opd_list),
            "total_target": total_target,
            "total_tertanam": total_planted,
            "overall_progress": min(overall_progress, 100)
        }
    }

# ============== KONTAK WHATSAPP ENDPOINTS ==============

@api_router.get("/kontak-whatsapp")
async def get_kontak_whatsapp():
    """Get WhatsApp contact settings (public endpoint)"""
    kontak = await db.kontak_whatsapp.find_one({}, {"_id": 0})
    if not kontak:
        return {"nomor_whatsapp": None, "pesan_default": None}
    return kontak

@api_router.post("/kontak-whatsapp", response_model=KontakWhatsAppResponse)
async def save_kontak_whatsapp(
    data: KontakWhatsAppCreate,
    current_user: dict = Depends(get_current_user)
):
    """Save WhatsApp contact settings (admin only)"""
    # Hapus kontak lama jika ada (hanya simpan 1 nomor aktif)
    await db.kontak_whatsapp.delete_many({})
    
    # Simpan kontak baru
    kontak_doc = {
        "nomor_whatsapp": data.nomor_whatsapp,
        "pesan_default": data.pesan_default or "",
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.kontak_whatsapp.insert_one(kontak_doc)
    
    return KontakWhatsAppResponse(
        nomor_whatsapp=data.nomor_whatsapp,
        pesan_default=data.pesan_default,
        updated_at=kontak_doc["updated_at"]
    )

# ============== EXPORT ENDPOINTS ==============

@api_router.get("/export/excel")
async def export_excel(current_user: dict = Depends(get_current_user)):
    partisipasi_list = await db.partisipasi.find({}, {"_id": 0}).to_list(10000)
    opd_list = await db.opd.find({}, {"_id": 0}).to_list(1000)
    opd_map = {o["id"]: o["nama"] for o in opd_list}
    
    # Tentukan jumlah maksimum lokasi
    max_lokasi = 1
    for p in partisipasi_list:
        lokasi_list = p.get("lokasi_list", [])
        if len(lokasi_list) > max_lokasi:
            max_lokasi = len(lokasi_list)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Data Partisipasi"
    
    # Header yang sesuai dengan format import
    # Format: Nama, NIP, Alamat, No. WhatsApp, OPD, Jumlah Pohon, Jenis Pohon, Sumber Bibit, Lokasi Tanam 1, Latitude 1, Longitude 1, ...
    headers = ["Nama", "NIP", "Alamat", "No. WhatsApp", "OPD", "Jumlah Pohon", "Jenis Pohon", "Sumber Bibit"]
    
    # Tambah kolom lokasi dinamis dengan Latitude dan Longitude terpisah
    for i in range(1, max_lokasi + 1):
        if max_lokasi == 1:
            headers.extend(["Lokasi Tanam", "Latitude", "Longitude"])
        else:
            headers.extend([f"Lokasi Tanam {i}", f"Latitude {i}", f"Longitude {i}"])
    
    ws.append(headers)
    
    for idx, p in enumerate(partisipasi_list, 1):
        row = [
            p.get("nama_lengkap", ""),
            p.get("nip", ""),
            p.get("alamat", ""),
            p.get("nomor_whatsapp", ""),
            opd_map.get(p.get("opd_id"), "Unknown"),
            p.get("jumlah_pohon", 0),
            p.get("jenis_pohon", ""),
            p.get("sumber_bibit", ""),
        ]
        
        # Tambahkan data lokasi dengan Latitude dan Longitude terpisah
        lokasi_list = p.get("lokasi_list", [])
        if not lokasi_list and p.get("lokasi_tanam"):
            # Fallback untuk data lama dengan single lokasi
            lokasi_list = [{"lokasi_tanam": p.get("lokasi_tanam", ""), "titik_lokasi": p.get("titik_lokasi", "")}]
        
        for i in range(max_lokasi):
            if i < len(lokasi_list):
                loc = lokasi_list[i]
                row.append(loc.get("lokasi_tanam", ""))
                # Parse titik_lokasi untuk mendapatkan lat/lng terpisah
                titik = loc.get("titik_lokasi", "")
                if titik and titik != "None" and "," in titik:
                    coords = titik.split(",")
                    row.append(coords[0].strip())  # Latitude
                    row.append(coords[1].strip() if len(coords) > 1 else "")  # Longitude
                else:
                    row.append("")  # Latitude
                    row.append("")  # Longitude
            else:
                row.append("")  # Lokasi Tanam
                row.append("")  # Latitude
                row.append("")  # Longitude
        
        ws.append(row)
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=data_partisipasi_agro_mopomulo.xlsx"}
    )

@api_router.get("/export/pdf")
async def export_pdf(current_user: dict = Depends(get_current_user)):
    partisipasi_list = await db.partisipasi.find({}, {"_id": 0}).to_list(10000)
    opd_list = await db.opd.find({}, {"_id": 0}).to_list(1000)
    opd_map = {o["id"]: o["nama"] for o in opd_list}
    
    # Tentukan jumlah maksimum lokasi (batasi 3 untuk PDF agar tidak terlalu lebar)
    max_lokasi = 1
    for p in partisipasi_list[:100]:
        lokasi_list = p.get("lokasi_list", [])
        if len(lokasi_list) > max_lokasi:
            max_lokasi = min(len(lokasi_list), 3)  # Batasi max 3 kolom lokasi untuk PDF
    
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
    
    elements = []
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], fontSize=16, spaceAfter=20, alignment=1)
    
    elements.append(Paragraph("Laporan Data Partisipasi Program Agro Mopomulo", title_style))
    elements.append(Paragraph(f"Kabupaten Gorontalo Utara - {datetime.now().strftime('%d %B %Y')}", styles['Normal']))
    elements.append(Spacer(1, 20))
    
    # Header dinamis
    headers = ["No", "Nama", "NIP", "OPD", "Pohon", "Jenis"]
    for i in range(1, max_lokasi + 1):
        if max_lokasi == 1:
            headers.append("Lokasi")
        else:
            headers.append(f"Lokasi {i}")
    
    data = [headers]
    
    for idx, p in enumerate(partisipasi_list[:100], 1):
        row = [
            str(idx),
            p.get("nama_lengkap", "")[:20],
            p.get("nip", "")[:15],
            opd_map.get(p.get("opd_id"), "")[:15],
            str(p.get("jumlah_pohon", 0)),
            p.get("jenis_pohon", "")[:12],
        ]
        
        # Tambahkan data lokasi
        lokasi_list = p.get("lokasi_list", [])
        if not lokasi_list and p.get("lokasi_tanam"):
            lokasi_list = [{"lokasi_tanam": p.get("lokasi_tanam", "")}]
        
        for i in range(max_lokasi):
            if i < len(lokasi_list):
                loc = lokasi_list[i]
                row.append(loc.get("lokasi_tanam", "")[:15])
            else:
                row.append("")
        
        data.append(row)
    
    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.Color(0.02, 0.59, 0.41)),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 7),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    
    elements.append(table)
    doc.build(elements)
    
    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": "attachment; filename=laporan_agro_mopomulo.pdf"}
    )

# ============== IMPORT ENDPOINTS ==============

@api_router.post("/import/excel")
async def import_excel(file: UploadFile = File(...), current_user: dict = Depends(get_current_user)):
    contents = await file.read()
    wb = load_workbook(filename=io.BytesIO(contents))
    ws = wb.active
    
    opd_list = await db.opd.find({}, {"_id": 0}).to_list(1000)
    opd_name_map = {o["nama"].lower(): o["id"] for o in opd_list}
    
    imported = 0
    errors = []
    
    # Baca header untuk menentukan format
    header_row = [str(cell.value).lower().strip() if cell.value else "" for cell in ws[1]]
    
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    for row_idx, row in enumerate(rows, 2):
        if not row or len(row) < 5:
            errors.append(f"Baris {row_idx}: Data tidak lengkap")
            continue
        
        # Skip row yang kosong
        if not any(row):
            continue
        
        try:
            # Format baru: Nama, NIP, Alamat, No. WhatsApp, OPD, Jumlah Pohon, Jenis Pohon, Sumber Bibit, Lokasi Tanam, Latitude, Longitude
            # Format lama: Nama, NIP, Email, OPD, Alamat, WA, Jumlah, Jenis, Lokasi
            
            # Cek apakah format baru (dengan header "alamat" di posisi 3)
            is_new_format = "alamat" in header_row and header_row.index("alamat") <= 3 if "alamat" in header_row else False
            
            if is_new_format or "latitude" in header_row or "sumber bibit" in header_row:
                # Format baru
                nama = row[0] if len(row) > 0 else ""
                nip = row[1] if len(row) > 1 else ""
                alamat = row[2] if len(row) > 2 else ""
                wa = row[3] if len(row) > 3 else ""
                opd_nama = row[4] if len(row) > 4 else ""
                jumlah = row[5] if len(row) > 5 else 0
                jenis = row[6] if len(row) > 6 else ""
                sumber_bibit = row[7] if len(row) > 7 else ""
                lokasi = row[8] if len(row) > 8 else ""
                latitude = row[9] if len(row) > 9 else ""
                longitude = row[10] if len(row) > 10 else ""
                email = ""
            else:
                # Format lama: Nama, NIP, Email, OPD, Alamat, WA, Jumlah, Jenis, Lokasi
                nama = row[0] if len(row) > 0 else ""
                nip = row[1] if len(row) > 1 else ""
                email = row[2] if len(row) > 2 else ""
                opd_nama = row[3] if len(row) > 3 else ""
                alamat = row[4] if len(row) > 4 else ""
                wa = row[5] if len(row) > 5 else ""
                jumlah = row[6] if len(row) > 6 else 0
                jenis = row[7] if len(row) > 7 else ""
                lokasi = row[8] if len(row) > 8 else ""
                sumber_bibit = ""
                latitude = ""
                longitude = ""
            
            if not nama:
                errors.append(f"Baris {row_idx}: Nama tidak boleh kosong")
                continue
            
            opd_id = opd_name_map.get(str(opd_nama).lower().strip()) if opd_nama else None
            if not opd_id:
                errors.append(f"Baris {row_idx}: OPD '{opd_nama}' tidak ditemukan")
                continue
            
            # Parse koordinat
            titik_lokasi = ""
            if latitude and longitude:
                titik_lokasi = f"{str(latitude).strip()}, {str(longitude).strip()}"
            
            # Buat lokasi_list
            lokasi_list = []
            if lokasi:
                lokasi_list.append({
                    "lokasi_tanam": str(lokasi).strip(),
                    "titik_lokasi": titik_lokasi,
                    "bukti_url": ""
                })
            
            # Cek apakah ada lokasi tambahan (Lokasi Tanam 2, Latitude 2, Longitude 2, dst)
            col_idx = 11  # Mulai dari kolom setelah Longitude pertama
            loc_num = 2
            while col_idx + 2 < len(row):
                lok = row[col_idx] if col_idx < len(row) else ""
                lat = row[col_idx + 1] if col_idx + 1 < len(row) else ""
                lng = row[col_idx + 2] if col_idx + 2 < len(row) else ""
                
                if lok:
                    titik = ""
                    if lat and lng:
                        titik = f"{str(lat).strip()}, {str(lng).strip()}"
                    lokasi_list.append({
                        "lokasi_tanam": str(lok).strip(),
                        "titik_lokasi": titik,
                        "bukti_url": ""
                    })
                
                col_idx += 3
                loc_num += 1
            
            partisipasi_id = str(uuid.uuid4())
            doc = {
                "id": partisipasi_id,
                "email": str(email).strip() if email else "",
                "nama_lengkap": str(nama).strip() if nama else "",
                "nip": str(nip).strip() if nip else "",
                "opd_id": opd_id,
                "alamat": str(alamat).strip() if alamat else "",
                "nomor_whatsapp": str(wa).strip() if wa else "",
                "jumlah_pohon": int(jumlah) if jumlah else 0,
                "jenis_pohon": str(jenis).strip() if jenis else "",
                "sumber_bibit": str(sumber_bibit).strip() if sumber_bibit else "",
                "lokasi_tanam": str(lokasi).strip() if lokasi else "",
                "titik_lokasi": titik_lokasi,
                "lokasi_list": lokasi_list,
                "created_at": datetime.now(timezone.utc).isoformat()
            }
            await db.partisipasi.insert_one(doc)
            imported += 1
        except Exception as e:
            errors.append(f"Baris {row_idx}: {str(e)}")
    
    return {"imported": imported, "errors": errors}

# ============== HEALTH CHECK ==============

# ============== DETEKSI GANDA (DUPLICATE DETECTION) ENDPOINTS ==============

class DuplicateGroupResponse(BaseModel):
    key_field: str  # "nama", "nip", or "nomor_whatsapp"
    key_value: str
    count: int
    participant_ids: List[str]

class DuplicateDetailItem(BaseModel):
    id: str
    nama_lengkap: str
    nip: Optional[str] = None
    nomor_whatsapp: Optional[str] = None
    opd_nama: Optional[str] = None
    jumlah_pohon: int
    jenis_pohon: str
    created_at: str

class MergeDuplicatesRequest(BaseModel):
    primary_id: str
    secondary_ids: List[str]

@api_router.get("/deteksi-ganda")
async def get_duplicates(
    field: str = "nama_lengkap",  # nama_lengkap, nip, nomor_whatsapp
    opd_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """
    Deteksi data duplikat berdasarkan field tertentu.
    Returns groups of duplicates with count > 1.
    """
    valid_fields = ["nama_lengkap", "nip", "nomor_whatsapp"]
    if field not in valid_fields:
        raise HTTPException(status_code=400, detail=f"Field harus salah satu dari: {', '.join(valid_fields)}")
    
    # Build match stage
    match_stage = {}
    if opd_id and opd_id != "all":
        match_stage["opd_id"] = opd_id
    
    # Exclude empty/null values for the field we're checking
    match_stage[field] = {"$ne": None, "$ne": ""}
    
    # Aggregation pipeline to find duplicates
    pipeline = [
        {"$match": match_stage},
        {
            "$group": {
                "_id": f"${field}",
                "count": {"$sum": 1},
                "participant_ids": {"$push": "$id"},
                "participants": {"$push": {
                    "id": "$id",
                    "nama_lengkap": "$nama_lengkap",
                    "nip": "$nip",
                    "nomor_whatsapp": "$nomor_whatsapp",
                    "opd_id": "$opd_id",
                    "jumlah_pohon": "$jumlah_pohon",
                    "jenis_pohon": "$jenis_pohon",
                    "created_at": "$created_at"
                }}
            }
        },
        {"$match": {"count": {"$gt": 1}}},  # Only groups with more than 1 entry
        {"$sort": {"count": -1}}  # Sort by count descending
    ]
    
    duplicates = await db.partisipasi.aggregate(pipeline).to_list(1000)
    
    # Get OPD map for enrichment
    opd_list = await db.opd.find({}, {"_id": 0}).to_list(1000)
    opd_map = {o["id"]: o["nama"] for o in opd_list}
    
    # Format response
    result = []
    for dup in duplicates:
        # Enrich participants with OPD names
        participants = []
        for p in dup.get("participants", []):
            p["opd_nama"] = opd_map.get(p.get("opd_id"), "Unknown")
            participants.append(p)
        
        result.append({
            "key_field": field,
            "key_value": dup["_id"],
            "count": dup["count"],
            "participant_ids": dup["participant_ids"],
            "participants": participants
        })
    
    return {
        "field": field,
        "total_groups": len(result),
        "total_duplicates": sum(d["count"] for d in result),
        "duplicates": result
    }

@api_router.delete("/deteksi-ganda/hapus")
async def delete_duplicates(
    ids: List[str],
    current_user: dict = Depends(get_current_user)
):
    """
    Hapus beberapa data partisipasi sekaligus (untuk menghapus duplikat).
    """
    if not ids:
        raise HTTPException(status_code=400, detail="Tidak ada ID yang diberikan")
    
    deleted_count = 0
    for pid in ids:
        result = await db.partisipasi.delete_one({"id": pid})
        deleted_count += result.deleted_count
    
    return {
        "success": True,
        "deleted_count": deleted_count,
        "message": f"Berhasil menghapus {deleted_count} data"
    }

@api_router.post("/deteksi-ganda/gabung")
async def merge_duplicates(
    request: MergeDuplicatesRequest,
    current_user: dict = Depends(get_current_user)
):
    """
    Gabungkan data duplikat menjadi satu.
    Data dari secondary_ids akan dihapus, dan jumlah pohon akan ditambahkan ke primary.
    lokasi_list dari semua data akan digabungkan.
    """
    if not primary_id or not secondary_ids:
        raise HTTPException(status_code=400, detail="primary_id dan secondary_ids diperlukan")
    
    # Get primary data
    primary = await db.partisipasi.find_one({"id": primary_id}, {"_id": 0})
    if not primary:
        raise HTTPException(status_code=404, detail="Data primer tidak ditemukan")
    
    # Get secondary data
    total_added_trees = 0
    merged_lokasi_list = list(primary.get("lokasi_list", []))
    
    # If primary has single lokasi, convert to list
    if not merged_lokasi_list and primary.get("lokasi_tanam"):
        merged_lokasi_list.append({
            "lokasi_tanam": primary.get("lokasi_tanam"),
            "titik_lokasi": primary.get("titik_lokasi"),
            "bukti_url": primary.get("bukti_url")
        })
    
    for sec_id in secondary_ids:
        secondary = await db.partisipasi.find_one({"id": sec_id}, {"_id": 0})
        if secondary:
            # Add trees
            total_added_trees += secondary.get("jumlah_pohon", 0)
            
            # Merge lokasi_list
            sec_lokasi_list = secondary.get("lokasi_list", [])
            if sec_lokasi_list:
                merged_lokasi_list.extend(sec_lokasi_list)
            elif secondary.get("lokasi_tanam"):
                merged_lokasi_list.append({
                    "lokasi_tanam": secondary.get("lokasi_tanam"),
                    "titik_lokasi": secondary.get("titik_lokasi"),
                    "bukti_url": secondary.get("bukti_url")
                })
            
            # Delete secondary
            await db.partisipasi.delete_one({"id": sec_id})
    
    # Update primary with merged data
    new_total_trees = primary.get("jumlah_pohon", 0) + total_added_trees
    
    await db.partisipasi.update_one(
        {"id": primary_id},
        {
            "$set": {
                "jumlah_pohon": new_total_trees,
                "lokasi_list": merged_lokasi_list
            }
        }
    )
    
    return {
        "success": True,
        "primary_id": primary_id,
        "merged_count": len(secondary_ids),
        "new_total_trees": new_total_trees,
        "total_locations": len(merged_lokasi_list),
        "message": f"Berhasil menggabungkan {len(secondary_ids)} data ke data primer"
    }

@api_router.get("/health")
async def health_check():
    return {"status": "healthy", "service": "Agro Mopomulo API"}

# Include router and middleware
app.include_router(api_router)

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
